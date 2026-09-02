#!/usr/bin/env python3
"""Official-source Atlanta Fed GDPNow + U.S. long-rate event watcher.

Triggers on a new Atlanta Fed GDPNow contribution-history observation.
Sources:
- Atlanta Fed GDPNow official Excel workbook (ContribHistory)
- Atlanta Fed official commentary page
- U.S. Treasury official daily par yield curve
- FRED DFII10 / T10YIE (Federal Reserve data)

Outputs:
- out/gdpnow_long_rates_status.md
- out/gdpnow_long_rates_alert.html (only when a new GDPNow observation appears)
- out/gdpnow_long_rates_pending_state.json
"""
from __future__ import annotations

import csv
import datetime as dt
import html
import io
import json
import pathlib
import re
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from html.parser import HTMLParser
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

KST = ZoneInfo("Asia/Seoul")
ROOT = pathlib.Path(__file__).resolve().parents[1]
OUT = ROOT / "out"
OUT.mkdir(parents=True, exist_ok=True)
STATE_PATH = ROOT / "data" / "gdpnow_long_rates_state.json"

GDP_XLSX = "https://www.atlantafed.org/-/media/Project/Atlanta/FRBA/Documents/cqer/researchcq/gdpnow/GDPTrackingModelDataAndForecasts.xlsx"
GDP_COMMENTARY = "https://www.atlantafed.org/research-and-data/data/gdpnow/current-and-past-gdpnow-commentaries"
TREASURY_XML_BASE = "https://home.treasury.gov/resource-center/data-chart-center/interest-rates/pages/xml"
FRED_CSV = "https://fred.stlouisfed.org/graph/fredgraph.csv"
USER_AGENT = "khs-watch-gdpnow-rates/1.0 (+https://github.com/qedgwangju-dot/khs-watch)"


@dataclass
class GdpRow:
    date: str
    release: str
    gdp: float
    pce: float | None
    equipment: float | None
    ipp: float | None
    nonres: float | None
    residential: float | None
    govt: float | None
    net_exports: float | None
    cipi: float


class TextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []

    def handle_data(self, data: str) -> None:
        text = re.sub(r"\s+", " ", data).strip()
        if text:
            self.parts.append(text)


def http_get(url: str, timeout: int = 35) -> bytes:
    req = urllib.request.Request(
        url,
        headers={"User-Agent": USER_AGENT, "Accept": "*/*", "Cache-Control": "no-cache"},
    )
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read()


def norm(v: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(v or "").lower())


def fnum(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("%", "").replace(",", "")
    if s in {"", ".", "NA", "N/A", "ND"}:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def fmt_date(v: Any) -> str:
    if isinstance(v, dt.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, dt.date):
        return v.isoformat()
    return str(v or "").strip()


def find_header_index(headers: list[str], names: list[str], contains: list[str] | None = None) -> int | None:
    wanted = {norm(x) for x in names}
    for i, h in enumerate(headers):
        if h in wanted:
            return i
    if contains:
        for i, h in enumerate(headers):
            if all(norm(x) in h for x in contains):
                return i
    return None


def fetch_contrib_rows() -> list[GdpRow]:
    raw = http_get(GDP_XLSX, timeout=50)
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet_name = next((n for n in wb.sheetnames if "contrib" in norm(n)), None)
    if not sheet_name:
        raise RuntimeError(f"ContribHistory sheet not found; sheets={wb.sheetnames}")
    ws = wb[sheet_name]

    header_row = None
    header_vals: list[str] = []
    for r in range(1, min(ws.max_row, 100) + 1):
        vals = [norm(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 24) + 1)]
        if "gdp" in vals and "cipi" in vals and any(v == "date" for v in vals):
            header_row = r
            header_vals = vals
            break
    if header_row is None:
        raise RuntimeError("Could not locate ContribHistory header row")

    idx_date = find_header_index(header_vals, ["Date"])
    idx_release = find_header_index(header_vals, ["Major Releases", "Major Release"], contains=["major", "release"])
    idx_gdp = find_header_index(header_vals, ["GDP"])
    idx_pce = find_header_index(header_vals, ["PCE"])
    idx_eq = find_header_index(header_vals, ["Equipment"], contains=["equip"])
    idx_ipp = find_header_index(header_vals, ["Intell. prop. prod.", "Intellectual property products"], contains=["intell"])
    idx_nonres = find_header_index(header_vals, ["Nonres. struct.", "Nonresidential structures"], contains=["nonres"])
    idx_res = find_header_index(header_vals, ["Resid. inves.", "Residential investment"], contains=["resid"])
    idx_gov = find_header_index(header_vals, ["Govt.", "Government"], contains=["gov"])
    idx_net = find_header_index(header_vals, ["Net exports"], contains=["net", "export"])
    idx_cipi = find_header_index(header_vals, ["CIPI"])

    required = {"date": idx_date, "gdp": idx_gdp, "cipi": idx_cipi}
    if any(v is None for v in required.values()):
        raise RuntimeError(f"Missing required ContribHistory columns: {required}; header={header_vals}")

    rows: list[GdpRow] = []
    for r in range(header_row + 1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 24) + 1)]
        gdp = fnum(vals[idx_gdp]) if idx_gdp is not None and idx_gdp < len(vals) else None
        cipi = fnum(vals[idx_cipi]) if idx_cipi is not None and idx_cipi < len(vals) else None
        date_v = vals[idx_date] if idx_date is not None and idx_date < len(vals) else None
        if gdp is None or cipi is None or date_v in (None, ""):
            continue

        def get(i: int | None) -> float | None:
            return fnum(vals[i]) if i is not None and i < len(vals) else None

        release = ""
        if idx_release is not None and idx_release < len(vals):
            release = str(vals[idx_release] or "").strip()
        rows.append(GdpRow(fmt_date(date_v), release, gdp, get(idx_pce), get(idx_eq), get(idx_ipp), get(idx_nonres), get(idx_res), get(idx_gov), get(idx_net), cipi))
    if not rows:
        raise RuntimeError("No usable ContribHistory observations")
    return rows


def fetch_commentary_meta() -> dict[str, str]:
    raw = http_get(GDP_COMMENTARY).decode("utf-8", errors="replace")
    p = TextExtractor(); p.feed(raw)
    text = " ".join(p.parts)
    out: dict[str, str] = {}
    m = re.search(r"The GDPNow model estimate for real GDP growth \(seasonally adjusted annual rate\) in the (first|second|third|fourth) quarter of (\d{4}) is ([0-9.]+) percent on ([A-Za-z]+ \d{1,2})", text, flags=re.I)
    if m:
        qmap = {"first": "Q1", "second": "Q2", "third": "Q3", "fourth": "Q4"}
        out.update(quarter=qmap[m.group(1).lower()], year=m.group(2), headline=m.group(3), commentary_date=m.group(4))
    n = re.search(r"The next GDPNow update is ([^.]+)\.", text, flags=re.I)
    if n:
        out["next_update"] = re.sub(r"\s+", " ", n.group(1)).strip()
    return out


def localname(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def fetch_treasury() -> dict[str, tuple[str, float]]:
    year = dt.datetime.now(KST).year
    params = urllib.parse.urlencode({"data": "daily_treasury_yield_curve", "field_tdr_date_value": str(year)})
    url = f"{TREASURY_XML_BASE}?{params}"
    root = ET.fromstring(http_get(url))
    records: list[dict[str, str]] = []
    for entry in root.iter():
        if localname(entry.tag) != "entry":
            continue
        rec: dict[str, str] = {}
        for node in entry.iter():
            if localname(node.tag) == "properties":
                for child in list(node):
                    rec[localname(child.tag)] = (child.text or "").strip()
                break
        if rec:
            records.append(rec)
    if not records:
        raise RuntimeError("Treasury XML returned no records")
    def dkey(rec: dict[str, str]) -> str:
        return (rec.get("NEW_DATE") or rec.get("QUOTE_DATE") or "")[:10]
    rec = sorted(records, key=dkey)[-1]
    d = dkey(rec)
    out: dict[str, tuple[str, float]] = {}
    for key, name in [("BC_10YEAR", "ust10"), ("BC_30YEAR", "ust30")]:
        v = fnum(rec.get(key))
        if v is None:
            raise RuntimeError(f"Treasury missing {key}")
        out[name] = (d, v)
    return out


def fetch_fred_latest(series: str) -> tuple[str, float]:
    url = f"{FRED_CSV}?{urllib.parse.urlencode({'id': series})}"
    text = http_get(url).decode("utf-8-sig", errors="replace")
    rows = []
    for row in csv.DictReader(io.StringIO(text)):
        d = (row.get("DATE") or row.get("observation_date") or "").strip()
        v = fnum(row.get(series))
        if d and v is not None:
            rows.append((d, v))
    if not rows:
        raise RuntimeError(f"FRED {series} returned no data")
    return rows[-1]


def load_state() -> dict[str, Any]:
    if not STATE_PATH.exists():
        return {}
    try:
        x = json.loads(STATE_PATH.read_text(encoding="utf-8"))
        return x if isinstance(x, dict) else {}
    except Exception:
        return {}


def main() -> int:
    now = dt.datetime.now(KST)
    errors: list[str] = []
    rows: list[GdpRow] = []
    try:
        rows = fetch_contrib_rows()
    except Exception as e:
        errors.append(f"Atlanta Fed Excel: {type(e).__name__}: {e}")
    meta: dict[str, str] = {}
    try:
        meta = fetch_commentary_meta()
    except Exception as e:
        errors.append(f"Atlanta Fed commentary: {type(e).__name__}: {e}")
    treasury: dict[str, tuple[str, float]] = {}
    try:
        treasury = fetch_treasury()
    except Exception as e:
        errors.append(f"U.S. Treasury: {type(e).__name__}: {e}")
    real10 = bei10 = None
    try:
        real10 = fetch_fred_latest("DFII10")
    except Exception as e:
        errors.append(f"FRED DFII10: {type(e).__name__}: {e}")
    try:
        bei10 = fetch_fred_latest("T10YIE")
    except Exception as e:
        errors.append(f"FRED T10YIE: {type(e).__name__}: {e}")

    if not rows:
        status = ["# 미국 GDPNow·장기금리 감시 상태", "", f"- 조회시각(KST): {now.isoformat(timespec='seconds')}", "- 상태: 확인 불가 — Atlanta Fed 공식 기여도 표를 읽지 못함"]
        if errors:
            status += ["", "## 오류"] + [f"- {x}" for x in errors]
        (OUT / "gdpnow_long_rates_status.md").write_text("\n".join(status) + "\n", encoding="utf-8")
        return 2

    latest = rows[-1]
    prev = rows[-2] if len(rows) >= 2 else None
    final_sales = latest.gdp - latest.cipi
    cipi_share = (latest.cipi / latest.gdp * 100.0) if abs(latest.gdp) > 0.05 else None
    fixed_components = [latest.equipment, latest.ipp, latest.nonres, latest.residential]
    private_final = latest.pce + sum(v for v in fixed_components if v is not None) if latest.pce is not None and all(v is not None for v in fixed_components) else None

    state = load_state()
    prior_key = str(state.get("observation_key") or "")
    obs_key = f"{latest.date}|{latest.gdp:.4f}|{latest.cipi:.4f}|{latest.release}"
    is_new = obs_key != prior_key
    pending = {"observation_key": obs_key, "latest": {"date": latest.date, "release": latest.release, "gdp": latest.gdp, "cipi": latest.cipi, "final_sales": final_sales}, "checked_at_kst": now.isoformat(timespec="seconds")}
    (OUT / "gdpnow_long_rates_pending_state.json").write_text(json.dumps(pending, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    status = ["# 미국 GDPNow·장기금리 감시 상태", "", f"- 조회시각(KST): {now.isoformat(timespec='seconds')}", f"- GDPNow 최신 관측: {latest.date} / {latest.gdp:.2f}% SAAR", f"- 재고 기여도(CIPI): {latest.cipi:+.2f}%p", f"- 재고 제외 최종판매(자체 계산): {final_sales:.2f}%", f"- 신규 업데이트: {'예' if is_new else '아니오'}"]
    if treasury:
        status.append(f"- UST 10Y/30Y: {treasury['ust10'][1]:.2f}% / {treasury['ust30'][1]:.2f}% ({treasury['ust10'][0]})")
    if errors:
        status += ["", "## 부분 확인 오류"] + [f"- {x}" for x in errors]
    (OUT / "gdpnow_long_rates_status.md").write_text("\n".join(status) + "\n", encoding="utf-8")

    alert_path = OUT / "gdpnow_long_rates_alert.html"
    if alert_path.exists(): alert_path.unlink()
    if not is_new: return 0

    quarter = f"{meta.get('year', '')} {meta.get('quarter', '')}".strip() or "현재 분기"
    delta = latest.gdp - prev.gdp if prev else None
    inv_delta = latest.cipi - prev.cipi if prev else None
    final_prev = (prev.gdp - prev.cipi) if prev else None
    final_delta = final_sales - final_prev if final_prev is not None else None

    verdicts: list[str] = []
    if cipi_share is not None and cipi_share >= 40:
        verdicts.append(f"헤드라인 성장의 약 {cipi_share:.0f}%가 재고 기여라 4분기 되돌림 위험이 큼")
    elif cipi_share is not None and cipi_share <= -20:
        verdicts.append("재고가 성장률을 크게 깎아 헤드라인보다 최종수요가 강할 수 있음")
    if final_sales >= 2.5:
        verdicts.append(f"재고를 제외한 최종판매도 {final_sales:.1f}%로 약하지 않음")
    if private_final is not None and private_final >= 3.0:
        verdicts.append(f"민간소비+고정투자 기여 합계가 약 {private_final:.1f}%p로 국내 민간수요가 강함")
    if latest.net_exports is not None and latest.net_exports <= -1.0:
        verdicts.append(f"순수출이 {latest.net_exports:.2f}%p를 깎아 강한 내수 일부가 수입으로 빠지는 구조")
    if treasury and (treasury["ust10"][1] >= 4.75 or treasury["ust30"][1] >= 5.0):
        verdicts.append("장기금리가 높은 구간이라 성장 호조가 성장주의 할인율 부담으로도 작동")
    if not verdicts:
        verdicts.append("헤드라인보다 재고·최종수요·순수출 구성을 함께 봐야 하는 구간")

    lines = ["📊 <b>미국 GDPNow·장기금리 업데이트</b>", "", f"조회: {now.strftime('%Y-%m-%d %H:%M:%S KST')}", f"분기: {html.escape(quarter)}", f"업데이트: {html.escape(latest.date)} · {html.escape(latest.release or '공식 기여도 갱신')}", "", "<b>핵심 숫자</b>", f"• GDPNow: <b>{latest.gdp:.2f}% SAAR</b>" + (f" (직전 {prev.gdp:.2f}%, {delta:+.2f}%p)" if prev else ""), f"• 재고 기여도(CIPI): <b>{latest.cipi:+.2f}%p</b>" + (f" (직전 대비 {inv_delta:+.2f}%p)" if inv_delta is not None else ""), f"• 재고 제외 최종판매: <b>{final_sales:.2f}%</b> = GDP {latest.gdp:.2f} - CIPI {latest.cipi:.2f}" + (f" (직전 대비 {final_delta:+.2f}%p)" if final_delta is not None else "")]
    if cipi_share is not None: lines.append(f"• 재고가 헤드라인 성장에서 차지하는 비중: 약 <b>{cipi_share:.1f}%</b>")
    if latest.pce is not None: lines.append(f"• 민간소비 기여도: {latest.pce:+.2f}%p")
    if private_final is not None: lines.append(f"• 민간소비+고정투자 기여 합계(자체 계산): 약 {private_final:+.2f}%p")
    if latest.net_exports is not None: lines.append(f"• 순수출 기여도: {latest.net_exports:+.2f}%p")
    lines += ["", "<b>장기금리 확인</b>"]
    if treasury: lines.append(f"• 미국 10Y {treasury['ust10'][1]:.2f}% · 30Y {treasury['ust30'][1]:.2f}% (미 재무부 {treasury['ust10'][0]})")
    else: lines.append("• 미국 10Y/30Y: 확인 불가")
    if real10: lines.append(f"• 미국 10Y 실질금리 {real10[1]:.2f}% (FRED {real10[0]})")
    if bei10: lines.append(f"• 미국 10Y 기대인플레이션 {bei10[1]:.2f}% (FRED {bei10[0]})")
    lines += ["", "<b>판정</b>"] + [f"• {html.escape(v)}" for v in verdicts]
    lines += ["", "<b>주의</b>", "• GDPNow는 Atlanta Fed의 실시간 모형 추정치이며 공식 FOMC 전망이 아님.", "• Final sales는 BEA 정의에 따라 GDP에서 민간재고 변화를 뺀 보조 계산.", "• 최대 반전 경로: 재고 축적이 실제 판매로 이어지지 않으면서 소비가 둔화하면 다음 분기 CIPI가 마이너스로 반전하고 장기금리도 빠르게 되돌릴 수 있음."]
    if meta.get("next_update"): lines.append(f"• 다음 Atlanta Fed 업데이트: {html.escape(meta['next_update'])}")
    lines += ["", "<b>공식 출처</b>", f"• <a href=\"{html.escape(GDP_COMMENTARY, quote=True)}\">Atlanta Fed GDPNow</a>", f"• <a href=\"{html.escape(GDP_XLSX, quote=True)}\">Atlanta Fed 기여도 원자료</a>", f"• <a href=\"{html.escape(TREASURY_XML_BASE, quote=True)}\">미 재무부 국채금리</a>", "• <a href=\"https://fred.stlouisfed.org/series/DFII10\">FRED 10Y 실질금리</a>", "• <a href=\"https://fred.stlouisfed.org/series/T10YIE\">FRED 10Y 기대인플레이션</a>"]
    alert_path.write_text("\n".join(lines).strip() + "\n", encoding="utf-8")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
